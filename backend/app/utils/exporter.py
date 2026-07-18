Input
import io
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from typing import Any, Iterable


def _autosize(ws):
    for col_idx, col_cells in enumerate(ws.columns, 1):
        try:
            letter = col_cells[0].column_letter
        except IndexError:
            continue
        max_len = 0
        for c in col_cells:
            if c.value is not None:
                max_len = max(max_len, len(str(c.value)))
        ws.column_dimensions[letter].width = min(50, max(10, max_len + 2))


def to_excel(headers: list[str], rows: Iterable[list[Any]], sheet_name: str = "Data",
             title: str | None = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    bold       = Font(bold=True, color="FFFFFF")
    blue_fill  = PatternFill("solid", fgColor="1E40AF")
    center     = Alignment(horizontal="center")

    row_offset = 1
    if title:
        ws.cell(row=1, column=1, value=title)
        ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="0F172A")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        row_offset = 3

    for c_idx, h in enumerate(headers, 1):
        c = ws.cell(row=row_offset, column=c_idx, value=h)
        c.font = bold; c.fill = blue_fill; c.alignment = center
    for r_idx, row in enumerate(rows, row_offset + 1):
        for c_idx, v in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=v)

    _autosize(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_csv(headers: list[str], rows: Iterable[list[Any]]) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(headers)
    for r in rows:
        w.writerow(r)
    return buf.getvalue().encode("utf-8")


def to_pdf_simple(title: str, headers: list[str], rows: Iterable[list[Any]]) -> bytes:
    """Minimal PDF generator without external dependencies.

    Writes a single-page A4 PDF with a monospaced table. Useful for
    report previews when jsPDF is not available server-side.
    """
    lines = [title, "", *["  ".join(str(c) for c in headers)]]
    for r in rows:
        lines.append("  ".join(str(c) for c in r))
    return _pdf_from_lines(lines)


def _pdf_from_lines(lines: list[str]) -> bytes:
    """Generate a minimal valid PDF (single page, Courier, A4)."""
    content_lines = ["BT", "/F1 9 Tf", "50 800 Td", "11 TL"]
    first = True
    for line in lines:
        escaped = line.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")
        if first:
            content_lines.append(f"({escaped}) Tj")
            first = False
        else:
            content_lines.append("T*")
            content_lines.append(f"({escaped}) Tj")
    content_lines.append("ET")
    stream = "\n".join(content_lines).encode("latin-1", errors="replace")

    objs: list[bytes] = []
    objs.append(b"<< /Type /Catalog /Pages 2 0 R >>")
    objs.append(b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>")
    objs.append(b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] "
                b"/Resources << /Font << /F1 5 0 R >> >> /Contents 4 0 R >>")
    objs.append(b"<< /Length " + str(len(stream)).encode() + b" >>\nstream\n" + stream + b"\nendstream")
    objs.append(b"<< /Type /Font /Subtype /Type1 /BaseFont /Courier >>")

    out = bytearray(b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n")
    offsets = [0]
    for i, o in enumerate(objs, 1):
        offsets.append(len(out))
        out.extend(f"{i} 0 obj\n".encode())
        out.extend(o)
        out.extend(b"\nendobj\n")
    xref_pos = len(out)
    out.extend(f"xref\n0 {len(objs) + 1}\n".encode())
    out.extend(b"0000000000 65535 f \n")
    for off in offsets[1:]:
        out.extend(f"{off:010d} 00000 n \n".encode())
    out.extend(f"trailer\n<< /Size {len(objs) + 1} /Root 1 0 R >>\nstartxref\n{xref_pos}\n%%EOF\n".encode())
    return bytes(out)
