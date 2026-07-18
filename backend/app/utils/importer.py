Input
import io
import logging
from typing import Any
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd

logger = logging.getLogger(__name__)


# Define headers + data-type hints + example rows per entity.
# Keys are stable identifiers used by the import endpoint.
ENTITY_TEMPLATES: dict[str, dict[str, Any]] = {
    "production_lines": {
        "headers": ["code", "name", "type", "capacity_per_hour",
                    "capacity_unit", "status", "changeover_minutes", "notes"],
        "types":   ["string", "string", "string", "number",
                    "string", "string", "number", "string"],
        "example": ["PL-01", "Packaging Line A", "discrete", 500,
                    "pcs", "active", 30, "Primary line"],
    },
    "machines": {
        "headers": ["code", "name", "line_id", "type", "capacity",
                    "capacity_unit", "criticality", "status", "notes"],
        "types":   ["string", "string", "number", "string", "number",
                    "string", "string", "string", "string"],
        "example": ["M-01", "Filler A", 1, "filler", 1000,
                    "pcs/h", "high", "active", ""],
    },
    "shifts": {
        "headers": ["name", "start_time", "end_time", "break_minutes",
                    "days_of_week", "headcount", "is_active"],
        "types":   ["string", "string", "string", "number",
                    "string", "number", "boolean"],
        "example": ["Morning", "06:00", "14:00", 30, "1,2,3,4,5", 5, True],
    },
    "products": {
        "headers": ["sku", "name", "category", "unit_of_measure",
                    "standard_cost", "selling_price", "min_order_qty",
                    "lead_time_days", "type"],
        "types":   ["string", "string", "string", "string",
                    "number", "number", "number", "number", "string"],
        "example": ["SKU-001", "Bottle 500ml", "Beverage", "pcs",
                    0.5, 1.25, 100, 3, "finished"],
    },
    "raw_materials": {
        "headers": ["code", "name", "category", "unit_of_measure",
                    "standard_cost", "safety_stock_qty", "reorder_point_qty",
                    "lead_time_days", "storage_conditions"],
        "types":   ["string", "string", "string", "string",
                    "number", "number", "number", "number", "string"],
        "example": ["RM-001", "PET Resin", "Plastic", "kg",
                    1.5, 100, 200, 7, "dry"],
    },
    "suppliers": {
        "headers": ["code", "name", "contact_name", "contact_email",
                    "contact_phone", "payment_terms_days", "rating", "status"],
        "types":   ["string", "string", "string", "string",
                    "string", "number", "number", "string"],
        "example": ["SUP-01", "Acme Resin Co", "John Smith",
                    "j@acme.com", "+2012345678", 30, 4, "active"],
    },
    "customers": {
        "headers": ["code", "name", "type", "priority",
                    "credit_limit", "payment_terms_days",
                    "contact_name", "contact_email", "contact_phone"],
        "types":   ["string", "string", "string", "number",
                    "number", "number", "string", "string", "string"],
        "example": ["CUST-01", "Hypermarket Co", "b2b", 1,
                    100000, 60, "Jane Doe", "jane@hm.com", "+20111222333"],
    },
    "warehouses": {
        "headers": ["code", "name", "type", "total_capacity",
                    "capacity_unit", "storage_conditions", "location"],
        "types":   ["string", "string", "string", "number",
                    "string", "string", "string"],
        "example": ["WH-01", "Main Warehouse", "raw_material", 5000,
                    "sqm", "ambient", "Cairo"],
    },
}


class DataImporter:
    def generate_template(self, entity: str) -> bytes:
        tpl = ENTITY_TEMPLATES.get(entity)
        if not tpl:
            raise ValueError(f"Unknown entity: {entity}")

        wb = Workbook()
        ws = wb.active
        ws.title = entity[:31]

        bold       = Font(bold=True, color="FFFFFF")
        blue_fill  = PatternFill("solid", fgColor="1E40AF")
        italic     = Font(italic=True, color="64748B")
        gray_fill  = PatternFill("solid", fgColor="F1F5F9")
        center     = Alignment(horizontal="center")

        ws.append(tpl["headers"])
        for c in ws[1]:
            c.font = bold; c.fill = blue_fill; c.alignment = center
        ws.append(tpl["types"])
        for c in ws[2]:
            c.font = italic; c.fill = gray_fill; c.alignment = center
        ws.append(tpl["example"])
        for c in ws[3]:
            c.fill = gray_fill

        for col_idx, header in enumerate(tpl["headers"], 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(15, len(header) + 2)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def import_and_validate(self, file_bytes: bytes, entity: str) -> dict:
        """Read an Excel/CSV upload and return a preview + validation result."""
        tpl = ENTITY_TEMPLATES.get(entity)
        if not tpl:
            raise ValueError(f"Unknown entity: {entity}")

        try:
            if file_bytes[:2] == b"PK":
                df = pd.read_excel(io.BytesIO(file_bytes), header=0)
            else:
                df = pd.read_csv(io.BytesIO(file_bytes), header=0)
        except Exception as e:
            return {"total": 0, "valid": 0, "errors": [{"row": 0, "field": "file", "message": str(e)}], "preview": [], "rows": []}

        df.columns = [str(c).strip().lower() for c in df.columns]
        expected = [h.lower() for h in tpl["headers"]]
        missing  = [c for c in expected if c not in df.columns]
        if missing:
            return {"total": len(df), "valid": 0,
                    "errors": [{"row": 0, "field": m, "message": f"Missing column: {m}"} for m in missing],
                    "preview": df.head(20).fillna("").to_dict(orient="records"), "rows": []}

        rows = df.fillna("").to_dict(orient="records")
        valid_rows: list = []
        errors:     list = []
        for i, row in enumerate(rows, start=2):
            errs = self._validate_row(row, tpl, i)
            if errs:
                errors.extend(errs)
            else:
                valid_rows.append(row)

        return {
            "total":  len(rows),
            "valid":  len(valid_rows),
            "errors": errors,
            "preview": rows[:20],
            "rows":    valid_rows,
        }

    def _validate_row(self, row: dict, tpl: dict, row_num: int) -> list:
        errs = []
        for header, dtype in zip(tpl["headers"], tpl["types"]):
            v = row.get(header.lower(), "")
            if dtype in ("number",) and v != "" and v is not None:
                try:
                    float(v)
                except (TypeError, ValueError):
                    errs.append({"row": row_num, "field": header, "message": f"Not a number: {v!r}"})
            if header in ("code", "name", "sku") and (v is None or v == ""):
                errs.append({"row": row_num, "field": header, "message": "Required field is empty"})
        return errs
